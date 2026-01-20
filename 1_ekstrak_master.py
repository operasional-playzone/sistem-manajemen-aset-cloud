import pandas as pd
import openpyxl

# --- KONFIGURASI MULTI-FILE ---
DAFTAR_FILE = [
    'Database_Aset_Lengkap.xlsx',       # File Jabodetabek
    'Database_Luar_Jabodetabek.xlsx'    # File Luar Jabodetabek
]

HEADER_KUNCI = "NAMA MESIN"

def cek_apakah_history(teks_header):
    if not isinstance(teks_header, str): return False
    kwd = teks_header.lower()
    keywords_history = ['mutasi', 'likuidasi', 'jual', 'spl', 'pindah', 'musnah']
    for word in keywords_history:
        if word in kwd: return True
    return False

def ekstrak_hanya_master():
    data_master_gabungan = []
    
    # --- LOOPING KE SETIAP FILE ---
    for nama_file in DAFTAR_FILE:
        print(f"\n📂 Membuka file: {nama_file}...")
        try:
            wb = openpyxl.load_workbook(nama_file, data_only=True)
        except FileNotFoundError:
            print(f"❌ File tidak ditemukan: {nama_file}, dilewati.")
            continue
            
        for nama_sheet in wb.sheetnames:
            print(f"   🔎 Scanning Sheet: {nama_sheet}...")
            ws = wb[nama_sheet]
            lokasi_toko = nama_sheet
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == HEADER_KUNCI:
                        anchor_row = cell.row
                        anchor_col = cell.column
                        
                        # Cek Header di Atasnya (Kategori)
                        header_atas = "Uncategorized"
                        try:
                            val = ws.cell(row=anchor_row-1, column=anchor_col-1).value
                            if not val: val = ws.cell(row=anchor_row-2, column=anchor_col-1).value
                            if val: header_atas = str(val)
                        except: pass
                        
                        # FILTER HISTORY (Skip jika ini tabel mutasi/likuidasi)
                        if cek_apakah_history(header_atas):
                            continue 
                        
                        kategori_bersih = header_atas.replace("KATEGORI", "").replace(":", "").strip()
                        
                        # SEDOT DATA
                        current_row = anchor_row + 1
                        while True:
                            nama_mesin = ws.cell(row=current_row, column=anchor_col).value
                            if not nama_mesin: break
                            
                            # --- MODIFIKASI: AMBIL DATA KOLOM LAIN ---
                            # Asumsi: Tanggal/Mesin Datang ada di SEBELAH KIRI Nama Mesin (-1)
                            # Jika ada di kanan, ganti jadi (anchor_col + 4) misalnya.
                            tgl_datang = ws.cell(row=current_row, column=anchor_col - 1).value
                            
                            harga      = ws.cell(row=current_row, column=anchor_col + 1).value
                            no_reg     = ws.cell(row=current_row, column=anchor_col + 2).value
                            no_sys     = ws.cell(row=current_row, column=anchor_col + 3).value
                            
                            # Bersihkan format tanggal jika perlu (Opsional)
                            if hasattr(tgl_datang, 'strftime'):
                                tgl_datang = tgl_datang.strftime('%Y-%m-%d')

                            data_master_gabungan.append({
                                'lokasi_toko': lokasi_toko,
                                'kategori': kategori_bersih,
                                'mesin_datang': tgl_datang,  # <--- KOLOM BARU
                                'nama_mesin': nama_mesin,
                                'harga_beli': harga,
                                'no_registrasi': no_reg,
                                'no_reg_system': no_sys,
                                'status': 'Aktif'
                            })
                            current_row += 1
                            
    return data_master_gabungan

# --- EKSEKUSI ---
list_master = ekstrak_hanya_master()
df = pd.DataFrame(list_master)

# Atur urutan kolom agar rapi saat di Excel
urutan_kolom = [
    'lokasi_toko', 'kategori', 'mesin_datang', 'nama_mesin', 
    'harga_beli', 'no_registrasi', 'no_reg_system', 'status'
]
# Pastikan hanya kolom yang ada yang diurutkan
df = df[[c for c in urutan_kolom if c in df.columns]]

print("\n=== HASIL MASTER ASET (GABUNGAN) ===")
print(f"✅ Total Aset Aktif: {len(df)} unit")
print(f"✅ Kolom 'mesin_datang' berhasil ditambahkan.")

output_file = '1_Master_Aset_Aktif.xlsx'
df.to_excel(output_file, index=False)
print(f"💾 File berhasil disimpan: {output_file}")