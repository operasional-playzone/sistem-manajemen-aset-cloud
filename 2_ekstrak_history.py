import pandas as pd
import openpyxl
import re

# --- KONFIGURASI MULTI-FILE ---
DAFTAR_FILE = [
    'Database_Aset_Lengkap.xlsx',       
    'Database_Luar_Jabodetabek.xlsx'    
]

# KEYWORDS UPPERCASE
KEYWORDS = ['MUTASI', 'LIKUIDASI', 'JUAL', 'SPL', 'MUSNAH', 'PINDAH', 'TARIK']

BULAN_INDO = {
    'januari': '01', 'februari': '02', 'maret': '03', 'april': '04',
    'mei': '05', 'juni': '06', 'juli': '07', 'agustus': '08',
    'september': '09', 'oktober': '10', 'november': '11', 'desember': '12',
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05',
    'jun': '06', 'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
}

def parse_header_info(teks):
    if not isinstance(teks, str): return None, None
    teks_upper = teks.upper()
    is_trigger = False
    for kw in KEYWORDS:
        if kw in teks_upper:
            is_trigger = True; break
    if not is_trigger: return None, None

    aksi = "History Lain"
    if "MUTASI" in teks_upper or "PINDAH" in teks_upper or "TARIK" in teks_upper: aksi = "Mutasi"
    elif "LIKUIDASI" in teks_upper or "JUAL" in teks_upper or "SPL" in teks_upper or "MUSNAH" in teks_upper: aksi = "Likuidasi"

    match = re.search(r'(\d{1,2})\s+([a-zA-Z]+)\s+(\d{4})', teks)
    tanggal_sql = None
    if match:
        tgl, bln_nama, thn = match.groups()
        bln_angka = BULAN_INDO.get(bln_nama.lower(), '01')
        tanggal_sql = f"{thn}-{bln_angka}-{tgl.zfill(2)}"
    return aksi, tanggal_sql

def cari_kategori_induk(ws, start_row, start_col):
    batas_atas = max(1, start_row - 50)
    for r in range(start_row - 1, batas_atas, -1):
        cek_sel_1 = ws.cell(row=r, column=start_col).value
        cek_sel_2 = ws.cell(row=r, column=start_col+1).value
        
        if cek_sel_1 == "NAMA MESIN" or cek_sel_2 == "NAMA MESIN":
            col_found = start_col if cek_sel_1 == "NAMA MESIN" else start_col+1
            try:
                kategori = ws.cell(row=r-1, column=col_found-1).value
                if not kategori: kategori = ws.cell(row=r-2, column=col_found-1).value
                if kategori: return str(kategori).replace("KATEGORI", "").replace(":", "").strip()
            except: pass
            return "Uncategorized (Header Found)"
    return "Uncategorized"

def scan_hanya_history():
    data_history_gabungan = []
    
    # --- LOOPING FILE ---
    for nama_file in DAFTAR_FILE:
        print(f"\n📂 Membuka file (Scan History): {nama_file}...")
        try:
            wb = openpyxl.load_workbook(nama_file, data_only=True)
        except FileNotFoundError:
            print(f"❌ File tidak ditemukan: {nama_file}, dilewati.")
            continue
            
        for nama_sheet in wb.sheetnames:
            print(f"   🔎 Scanning Sheet: {nama_sheet}...")
            ws = wb[nama_sheet]
            
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    aksi, tanggal = parse_header_info(val)
                    
                    if aksi: 
                        anchor_row = cell.row
                        anchor_col = cell.column
                        keterangan_full = str(val)
                        kategori_ditemukan = cari_kategori_induk(ws, anchor_row, anchor_col)
                        
                        current_row = anchor_row + 1
                        while True:
                            # Logika offset kolom +1
                            nama_mesin = ws.cell(row=current_row, column=anchor_col + 1).value
                            if not nama_mesin: break
                            
                            harga   = ws.cell(row=current_row, column=anchor_col + 2).value
                            no_reg  = ws.cell(row=current_row, column=anchor_col + 3).value
                            no_sys  = ws.cell(row=current_row, column=anchor_col + 4).value
                            
                            data_history_gabungan.append({
                                'lokasi_asal': nama_sheet,
                                'kategori': kategori_ditemukan,
                                'jenis_aksi': aksi,
                                'tanggal': tanggal,
                                'nama_mesin': nama_mesin,
                                'harga_beli': harga,
                                'no_registrasi': no_reg,
                                'no_reg_system': no_sys,
                                'keterangan': keterangan_full 
                            })
                            current_row += 1 

    return data_history_gabungan

# --- EKSEKUSI ---
list_hist = scan_hanya_history()
df = pd.DataFrame(list_hist)

print("\n=== HASIL HISTORY LOG GABUNGAN ===")
print(f"✅ Total Data History: {len(df)} baris")

output_file = '2_Riwayat_Log_Fix.xlsx'
df.to_excel(output_file, index=False)
print(f"💾 File berhasil disimpan: {output_file}")